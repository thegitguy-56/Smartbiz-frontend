"""
SmartBiz E2E Test Report Generator
===================================
Parses pytest JUnit XML results and generates a formatted .xlsx
report with 5 sheets matching the reference report format:
  1. Summary       – Overall test suite statistics
  2. Passed Tests  – All passing test cases with timing
  3. Failed Tests  – All failing test cases with error details
  4. Test Details  – Complete list of every test case
  5. Execution Log – Timestamped log of test execution events
"""

import os
import sys
import json
import xml.etree.ElementTree as ET
from datetime import datetime, timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Styling constants
# ---------------------------------------------------------------------------
HEADER_FONT = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
PASS_FONT = Font(name="Calibri", bold=True, color="006100")
FAIL_FONT = Font(name="Calibri", bold=True, color="9C0006")
NORMAL_FONT = Font(name="Calibri", size=11)
BOLD_FONT = Font(name="Calibri", bold=True, size=11)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")


# ---------------------------------------------------------------------------
# Category mapping – derive category from test class/file name
# ---------------------------------------------------------------------------
CATEGORY_MAP = {
    "test_01_onboarding": "Onboarding Page",
    "test_02_login": "Login Page",
    "test_03_register": "Registration Page",
    "test_04_dashboard": "Dashboard",
    "test_05_inventory": "Inventory",
    "test_06_add_product": "Add Product",
    "test_07_billing": "Billing / Invoice",
    "test_08_customers": "Customers",
    "test_09_reports": "Reports",
    "test_10_settings": "Settings",
    "test_11_navigation": "Navigation & Routing",
    "test_12_security": "Security & Vulnerability",
}


def get_category(classname: str) -> str:
    """Derive a human-readable category from the test classname."""
    for key, value in CATEGORY_MAP.items():
        if key in classname:
            return value
    return "General"


# ---------------------------------------------------------------------------
# Parse JUnit XML
# ---------------------------------------------------------------------------
def parse_junit_xml(xml_path: str) -> dict:
    """
    Parse pytest JUnit XML and return structured test results.
    Returns a dict with keys: suite_name, total, passed, failed,
    duration, start_time, end_time, tests (list of dicts).
    """
    tree = ET.parse(xml_path)
    root = tree.getroot()

    tests = []
    total_time = 0.0

    # Handle both <testsuites> and <testsuite> root elements
    if root.tag == "testsuites":
        suites = root.findall("testsuite")
    else:
        suites = [root]

    for suite in suites:
        for testcase in suite.findall("testcase"):
            name = testcase.get("name", "unknown")
            classname = testcase.get("classname", "")
            time_taken = float(testcase.get("time", "0"))
            total_time += time_taken

            failure = testcase.find("failure")
            error = testcase.find("error")
            skipped = testcase.find("skipped")

            if failure is not None:
                status = "FAILED"
                error_msg = failure.get("message", "")
                error_detail = failure.text or error_msg
            elif error is not None:
                status = "FAILED"
                error_msg = error.get("message", "")
                error_detail = error.text or error_msg
            elif skipped is not None:
                status = "SKIPPED"
                error_msg = skipped.get("message", "Skipped")
                error_detail = error_msg
            else:
                status = "PASSED"
                error_msg = "None \u2013 test passed successfully."
                error_detail = error_msg

            tests.append({
                "name": name,
                "classname": classname,
                "category": get_category(classname),
                "status": status,
                "time": round(time_taken, 2),
                "error": error_msg,
                "error_detail": error_detail,
            })

    passed = [t for t in tests if t["status"] == "PASSED"]
    failed = [t for t in tests if t["status"] == "FAILED"]

    now = datetime.now(timezone.utc)
    pass_rate = (len(passed) / len(tests) * 100) if tests else 0

    return {
        "suite_name": "SmartBiz Web App \u2013 Full E2E Workflow",
        "total": len(tests),
        "passed": len(passed),
        "failed": len(failed),
        "pass_rate": round(pass_rate, 2),
        "duration": round(total_time, 2),
        "start_time": now.isoformat(),
        "end_time": now.isoformat(),
        "tests": tests,
        "passed_tests": passed,
        "failed_tests": failed,
    }


# ---------------------------------------------------------------------------
# Parse execution log
# ---------------------------------------------------------------------------
def parse_execution_log(log_path: str) -> list:
    """Parse the JSONL execution log file."""
    entries = []
    if not os.path.exists(log_path):
        return entries

    with open(log_path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                data = json.loads(line)
                entries.append(data)
            except json.JSONDecodeError:
                # Try to parse as a wrapped JSON string
                try:
                    inner = json.loads(json.loads(f'"{line}"'))
                    entries.append(inner)
                except Exception:
                    entries.append({
                        "timestamp": datetime.now(timezone.utc).isoformat(),
                        "level": "INFO",
                        "message": line,
                    })
    return entries


# ---------------------------------------------------------------------------
# Style helper
# ---------------------------------------------------------------------------
def style_header_row(ws, row_num: int, num_cols: int):
    """Apply header styling to a row."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def auto_width(ws, min_width=12, max_width=80):
    """Auto-adjust column widths based on content."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                cell_len = len(str(cell.value or ""))
                max_len = max(max_len, cell_len)
            except Exception:
                pass
        adjusted = min(max(max_len + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted


# ---------------------------------------------------------------------------
# Generate XLSX report
# ---------------------------------------------------------------------------
def generate_report(results: dict, log_entries: list, output_path: str):
    """Generate the full .xlsx test report."""
    wb = Workbook()

    # ===== Sheet 1: Summary =====
    ws_summary = wb.active
    ws_summary.title = "Summary"
    headers = [
        "Test Suite", "Total Tests", "Passed", "Failed",
        "Pass Rate %", "Duration (sec)", "Start Time", "End Time",
    ]
    ws_summary.append(headers)
    style_header_row(ws_summary, 1, len(headers))
    ws_summary.append([
        results["suite_name"],
        results["total"],
        results["passed"],
        results["failed"],
        results["pass_rate"],
        results["duration"],
        results["start_time"],
        results["end_time"],
    ])
    # Style the data row
    for col in range(1, len(headers) + 1):
        cell = ws_summary.cell(row=2, column=col)
        cell.font = NORMAL_FONT
        cell.border = THIN_BORDER
    auto_width(ws_summary)

    # ===== Sheet 2: Passed Tests =====
    ws_passed = wb.create_sheet("Passed Tests")
    p_headers = ["No.", "Category", "Test Name", "Time (sec)", "Status"]
    ws_passed.append(p_headers)
    style_header_row(ws_passed, 1, len(p_headers))

    for i, t in enumerate(results["passed_tests"], 1):
        ws_passed.append([i, t["category"], t["name"], t["time"], t["status"]])
        row = ws_passed.max_row
        for col in range(1, len(p_headers) + 1):
            cell = ws_passed.cell(row=row, column=col)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
        # Color the Status cell green
        ws_passed.cell(row=row, column=5).fill = PASS_FILL
        ws_passed.cell(row=row, column=5).font = PASS_FONT
    auto_width(ws_passed)

    # ===== Sheet 3: Failed Tests =====
    ws_failed = wb.create_sheet("Failed Tests")
    f_headers = ["No.", "Category", "Test Name", "Error", "Status", "Timestamp"]
    ws_failed.append(f_headers)
    style_header_row(ws_failed, 1, len(f_headers))

    timestamp_str = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
    for i, t in enumerate(results["failed_tests"], 1):
        ws_failed.append([
            i, t["category"], t["name"],
            t["error_detail"][:500],  # Truncate long errors
            t["status"], timestamp_str,
        ])
        row = ws_failed.max_row
        for col in range(1, len(f_headers) + 1):
            cell = ws_failed.cell(row=row, column=col)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            cell.alignment = WRAP_ALIGNMENT
        # Color the Status cell red
        ws_failed.cell(row=row, column=5).fill = FAIL_FILL
        ws_failed.cell(row=row, column=5).font = FAIL_FONT
    auto_width(ws_failed)

    # ===== Sheet 4: Test Details =====
    ws_details = wb.create_sheet("Test Details")
    d_headers = ["No.", "Category", "Test Name", "Status", "Error Details"]
    ws_details.append(d_headers)
    style_header_row(ws_details, 1, len(d_headers))

    for i, t in enumerate(results["tests"], 1):
        ws_details.append([
            i, t["category"], t["name"], t["status"],
            t["error_detail"][:500],
        ])
        row = ws_details.max_row
        for col in range(1, len(d_headers) + 1):
            cell = ws_details.cell(row=row, column=col)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            cell.alignment = WRAP_ALIGNMENT
        # Color status cell
        status_cell = ws_details.cell(row=row, column=4)
        if t["status"] == "PASSED":
            status_cell.fill = PASS_FILL
            status_cell.font = PASS_FONT
        elif t["status"] == "FAILED":
            status_cell.fill = FAIL_FILL
            status_cell.font = FAIL_FONT
    auto_width(ws_details)

    # ===== Sheet 5: Execution Log =====
    ws_log = wb.create_sheet("Execution Log")
    l_headers = ["Timestamp", "Level", "Message"]
    ws_log.append(l_headers)
    style_header_row(ws_log, 1, len(l_headers))

    for entry in log_entries:
        ws_log.append([
            entry.get("timestamp", ""),
            entry.get("level", "INFO"),
            entry.get("message", ""),
        ])
        row = ws_log.max_row
        for col in range(1, len(l_headers) + 1):
            cell = ws_log.cell(row=row, column=col)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
    auto_width(ws_log)

    # Save workbook
    wb.save(output_path)
    print(f"\n{'='*60}")
    print(f"  XLSX Report Generated: {output_path}")
    print(f"  Total Tests:  {results['total']}")
    print(f"  Passed:       {results['passed']}")
    print(f"  Failed:       {results['failed']}")
    print(f"  Pass Rate:    {results['pass_rate']}%")
    print(f"{'='*60}\n")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    """Entry point: parse XML results and generate XLSX report."""
    # Determine paths
    script_dir = os.path.dirname(os.path.abspath(__file__))
    xml_path = os.path.join(script_dir, "results.xml")
    log_path = os.path.join(script_dir, "execution_log.jsonl")

    # Allow overriding via command-line args
    if len(sys.argv) > 1:
        xml_path = sys.argv[1]
    if len(sys.argv) > 2:
        log_path = sys.argv[2]

    if not os.path.exists(xml_path):
        print(f"ERROR: JUnit XML not found at {xml_path}")
        print("Run tests first: pytest e2e/ -v --junitxml=results.xml")
        sys.exit(1)

    # Parse results
    results = parse_junit_xml(xml_path)
    log_entries = parse_execution_log(log_path)

    # Generate output filename with timestamp
    timestamp = datetime.now().strftime("%Y-%m-%dT%H-%M-%S")
    output_filename = f"E2E_Test_Report_SmartBiz_{timestamp}.xlsx"
    output_path = os.path.join(script_dir, output_filename)

    # Generate report
    generate_report(results, log_entries, output_path)

    return output_path


if __name__ == "__main__":
    main()
